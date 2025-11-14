"""
Excel Autofiller Module
Handles automated filling of cash sheet Excel workbooks with parsed sales data.
"""

from openpyxl import load_workbook
import traceback
from config import FILL_COL_MAP, CHECKING_COL_MAP


class ExcelAutofiller:
    """
    Automates the process of filling cash sheet Excel files with sales data.

    Attributes:
        xl_path (str): Path to the Excel workbook file
        location (str): Location name to find and fill in the worksheet
        week_day (str): Name of the worksheet tab (day of week)
        start_row (int): Starting row for location search (default: 4)
        row (int): Current row being processed
        wb: Openpyxl workbook object
        ws: Openpyxl worksheet object
    """

    def __init__(self, xl_path, location, week_day):
        """
        Initialize the ExcelAutofiller.

        Args:
            xl_path (str): Full path to the Excel workbook
            location (str): Location name to search for in the workbook
            week_day (str): Worksheet name (e.g., 'Monday', 'Tuesday')
        """
        self.xl_path = xl_path
        self.location = location
        self.week_day = week_day
        self.start_row = 4  # Data starts at row 4 (after headers)
        self.row = 0
        self.wb = None
        self.ws = None

    def open_workbook(self):
        """
        Open and load the Excel workbook and specified worksheet.

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            self.wb = load_workbook(self.xl_path)

            # Check if the worksheet exists
            if self.week_day not in self.wb.sheetnames:
                print(f"  ❌ Worksheet '{self.week_day}' not found in workbook")
                print(
                    f"     Available sheets: {', '.join(self.wb.sheetnames)}")
                return False

            self.ws = self.wb[self.week_day]
            return True

        except FileNotFoundError:
            print(f"  ❌ File not found: {self.xl_path}")
            return False

        except PermissionError:
            print(f"  ❌ Permission denied: {self.xl_path}")
            print("     Please close the file if it's currently open")
            return False

        except Exception as e:
            print(f"  ❌ Unexpected error opening workbook: {e}")
            return False

    def find_row(self):
        """
        Search for the location name in the worksheet and set the current row.

        Returns:
            bool: True if location found, False otherwise
        """
        location_col = FILL_COL_MAP.get("location")

        if location_col is None:
            print("  ❌ 'location' column not defined in FILL_COL_MAP")
            return False

        # Search through rows starting from start_row
        for r in range(self.start_row, self.ws.max_row + 1):
            cell_value = self.ws.cell(r, location_col).value

            # Check if cell has value and matches location (case-insensitive, stripped)
            if cell_value and cell_value.strip().lower() == self.location.lower():
                self.row = r
                print(
                    f"  ✓ Found location '{self.location}' at row {self.row}")
                return True

        # Location not found
        print(
            f"  ❌ Location '{self.location}' not found in worksheet '{self.week_day}'")
        print(f"     Searched rows {self.start_row} to {self.ws.max_row}")
        return False

    def checking_tenders(self):
        """
        Verify that tender calculations are correct by checking the 'over/short' column.

        Returns:
            bool: True if tenders balance correctly (over/short is 0 or None), False otherwise
        """
        over_col = CHECKING_COL_MAP.get("over")

        if over_col is None:
            print(
                "  ⚠️  'over' column not defined in CHECKING_COL_MAP - skipping validation")
            return True

        over_value = self.ws.cell(self.row, over_col).value

        # Check if there's a discrepancy
        if over_value is not None:
            try:
                over_amount = float(over_value)
                if over_amount != 0:
                    print(
                        f"  ⚠️  Tender discrepancy detected: ${over_amount:.2f}")
                    return False
            except (ValueError, TypeError):
                print(
                    f"  ⚠️  Invalid value in over/short column: {over_value}")
                return False

        return True

    def filling(self, parser):
        """
        Fill the worksheet with data from the parser dictionary.

        Args:
            parser (dict): Dictionary containing parsed data with keys:
                - date: Report date
                - count: Guest count
                - total_sales: Total sales amount
                - tax: Tax amount
                - tenders: Dictionary of tender types and amounts

        Returns:
            bool: True if filling successful, False otherwise
        """
        try:
            # Step 1: Find the correct row for this location
            if not self.find_row():
                return False

            # Tax is typically on the row below the main data row
            tax_row = self.row + 1

            # Step 2: Fill basic performance metrics
            date_col = FILL_COL_MAP.get("date")
            if date_col:
                self.ws.cell(1, date_col).value = parser.get("date")

            count_col = FILL_COL_MAP.get("count")
            if count_col:
                self.ws.cell(self.row, count_col).value = parser.get("count")

            total_sales_col = FILL_COL_MAP.get("total_sales")
            if total_sales_col:
                self.ws.cell(self.row, total_sales_col).value = parser.get(
                    "total_sales")

            tax_col = FILL_COL_MAP.get("tax")
            if tax_col:
                self.ws.cell(tax_row, tax_col).value = parser.get("tax")

            # Step 3: Fill tender amounts
            tenders = parser.get("tenders", {})
            unmatched_tenders = []

            for tender_name, amount in tenders.items():
                if tender_name not in FILL_COL_MAP:
                    unmatched_tenders.append(tender_name)
                    continue

                col = FILL_COL_MAP[tender_name]

                # Only fill non-zero amounts; clear zero amounts
                if amount > 0:
                    self.ws.cell(self.row, col).value = amount
                elif amount == 0:
                    self.ws.cell(self.row, col).value = None

            # Report any unmatched tenders
            if unmatched_tenders:
                print(
                    f"  ⚠️  Unmatched tenders (not filled): {', '.join(unmatched_tenders)}")

            print(f"  ✓ {self.location} filled successfully")
            return True

        except KeyError as e:
            print(f"  ❌ Missing required data in parser: {e}")
            return False

        except Exception as e:
            print(f"  ❌ Error filling {self.location}: {e}")
            traceback.print_exc()
            return False

    def save(self):
        """
        Save the workbook and validate the filled data.

        Returns:
            bool: True if save successful and data validates correctly, False otherwise
        """
        try:
            # Step 1: Save the workbook
            self.wb.save(self.xl_path)
            print(f"  💾 Saved: {self.location} casheet")

            # Step 2: Reload workbook with calculated formulas (data_only=True)
            self.wb = load_workbook(self.xl_path, data_only=True)
            self.ws = self.wb[self.week_day]

            # Step 3: Validate tender calculations
            is_correct = self.checking_tenders()

            if is_correct:
                print(
                    f"  ✅ {self.location} validated - all tenders balance correctly")
            else:
                print(
                    f"  ⚠️  {self.location} validation warning - check over/short column")

            return is_correct

        except PermissionError:
            print(f"  ❌ Cannot save: File is open in another program")
            return False

        except Exception as e:
            print(f"  ❌ Save error: {e}")
            return False

    def close(self):
        """
        Properly close the workbook to free resources.
        """
        if self.wb:
            self.wb.close()
            self.wb = None
            self.ws = None

    def __enter__(self):
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - ensures workbook is closed."""
        self.close()
